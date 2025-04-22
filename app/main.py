import subprocess
import datetime
import os
import shutil
import yaml
from openpyxl import Workbook, load_workbook
from datetime import datetime
from utils.classes import AppFiles
from utils.cmd_util import CmdUtil
from core.svn_tracker import SvnTracker

"""
# SVN Status를 받아서 풀네임 반환
def svn_status_fullname(code: str) -> str:
    status_mapping = {
        'A': 'Added',
        'C': 'Conflicted',
        'D': 'Deleted',
        'I': 'Ignored',
        'M': 'Modified',
        'R': 'Replaced',
        'X': 'External',
        '?': 'Unversioned',
        '!': 'Missing',
        '~': 'Obstructed',
    }

    return status_mapping.get(code.upper(), 'Unknown')
"""
"""
def run_svn_commit(repo_path, commit_list, commit_message):
    try:
        commit_cmd = ["svn", "commit", "-m", commit_message]

        for _, file_name in commit_list:
            commit_cmd.append(file_name)

        subprocess.run(commit_cmd, cwd=repo_path, check=True)
    except subprocess.CalledProcessError as e:
        print("SVN 커밋에 실패하였습니다.")
        print(e)
        exit(1)

def get_last_author(file_path, repo_path):
    try:
        result = subprocess.run(["svn", "log", "-l", "1", file_path], cwd=repo_path, capture_output=True, text=True)
        lines = result.stdout.splitlines()
        if len(lines) >= 2:
            parts = lines[1].split('|')
            if len(parts) >= 2:
                return parts[1].strip()
    except Exception:
        pass
    return "Unknown"

def copy_changed_files(commit_list, src_root, dest_root):
    # dest 경로가 없으면 폴더 생성
    mkdir(dest_root)

    for _, file_path in commit_list:
        full_src_path = os.path.join(src_root, file_path)
        full_dest_path = os.path.join(dest_root, file_path)
        mkdir(os.path.dirname(full_dest_path))

        if os.path.isfile(full_src_path):
            shutil.copy2(full_src_path, full_dest_path)
            print(f"- from {full_src_path}")
            print(f"- to {full_dest_path}")

def write_to_excel(changes, repo_path, excel_path):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 파일이 없으면 새로 생성
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Author", "Change Type", "File Path"])
    else:
        wb = load_workbook(excel_path)
        ws = wb.active

    for change_type, file_path in changes:
        author = get_last_author(file_path, repo_path)
        ws.append([timestamp, author, change_type, file_path])

    wb.save(excel_path)

def export_existing_files(remote_svn_path, before_path, commit_list):
    failed_list = []

    for type, file_path in commit_list:
        remote_file_path = remote_svn_path + "/" + file_path
        local_save_path = before_path + "/" + file_path

        try:
            mkdir(os.path.dirname(local_save_path))
            subprocess.run(["svn", "export", remote_file_path, local_save_path], check=True)
        except subprocess.CalledProcessError as e:
            print("😭😭 {remote_file_path} 파일 다운로드 실패")
            print(e)
            failed_list.append((svn_status_fullname(type), file_path))
    
    return failed_list

def get_svn_status(repo_path):
    result = subprocess.run(["svn", "status"], cwd=repo_path, capture_output=True, text=True)
    return result.stdout

def can_append_path(path, repo_path) -> bool:
    result = subprocess.run(["svn", "status", "-u", path], cwd=repo_path, capture_output=True, text=True)
    print(result)
    return True

def parse_status_output(output):
    print(output)
    # A	추가됨 (Added)
    # C	충돌 (Conflicted)
    # D	삭제됨 (Deleted)
    # I	무시됨 (Ignored)
    # M	수정됨 (Modified)
    # R	교체됨 (Replaced)
    # X	외부 참조 (eXternals definition)
    # ?	버전 관리되지 않음 (Unversioned)
    # !	누락됨 (item missing, may be deleted)
    # ~	타입 변경됨 (Type changed)

    # 복합 상태가 나오는 경우
    # A +	(A: Added, +: with history)	파일을 추가했는데, 기존 파일 복사해서 만든 경우
    # R +	(R: Replaced, +: with history)	기존 파일 삭제 후 새 파일을 추가한 경우, 히스토리 있음
    # R C	(R: Replaced, C: Conflict)	파일을 교체했는데 충돌도 남
    # M C	(M: Modified, C: Conflict)	수정 중 충돌 발생
    # A C	(A: Added, C: Conflict)	추가한 파일에 충돌 발생 (드물지만 가능)

    BASE_STATUS = ['A', 'C', 'D', 'I', 'M', 'R', 'X', '?', '!', '~', 'A']
    changes = []

    for line in output.splitlines():
        if not line.strip():
            continue

        status_code = line[:7].strip()
        path = line[7:].strip()

        # 기본 상태인 경우
        if status_code in BASE_STATUS:
            # 충돌 파일 제외
            if status_code == 'C':
                print("- [Exclude] 충돌 파일 제외")
                print(line)
                print()
                continue

            # Unversioned file 제외
            if status_code == '?':
                print(f"- [Exclude] {path}  --  Unversioned File")
                continue
            
            print(f"- [Include] {path}")
            path = path.replace("\\", "/")
            status_code = svn_status_fullname(status_code)
            changes.append((status_code, path))

    return changes

def main(repo_path, remote_path, ftp_list) -> None:
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest_path = os.path.join(backup_path, now)
    before_path = os.path.join(backup_path, now, "before")
    after_path = os.path.join(backup_path, now, "after")

    for path in [before_path, after_path]:
        mkdir(path)

    print("=" * 110)
    print("\n\n[ SVN STATUS ]") 
    print(f"👌 Target Repository Name: {repo_name}")
    print(f"📁 Repository Path: {repo_path}")
    print(f"📁 Backup Path: {dest_path}")
    print("\n🔍 Checking for changes...")
    status_output = get_svn_status(repo_path)

    changes = parse_status_output(status_output)
    print("\n")

    if not changes:
        print("✅ 변경된 파일이 없습니다.")
        print("프로그램을 종료합니다.")
        press_input()
        return

    commit_list = []
    print("[ 커밋할 대상 파일을 선택해주세요. ]")

    for type, file_path in changes:
        ans = input(f"({type}) {file_path} (Y/N): ")
        if ans.upper() == "Y":
            commit_list.append((type, file_path))

    if len(commit_list) == 0:
        print("✅ 선택된 커밋 대상 파일이 없습니다.")
        print("프로그램을 종료합니다.")
        press_input()
        return

    print("\n\n[ 선택된 커밋 대상 파일 ]")
    for type, file_path in commit_list:
        print(f"- ({type}) {file_path}")

    rst = input("\n\커밋을 진행할까요? (Y/N): ")

    if rst.upper() != "Y":
        print("프로그램을 종료합니다.")
        press_input()
        return

    while True:
        commit_message = input("커밋 메시지를 입력해주세요: ")
        if len(commit_message) > 0:
            break
        else:
            print("[경고] 커밋 메시지는 최소 1글자 이상 입력해야 합니다.")

    print("\n\n[ Remote 서버의 파일을 백업 중입니다... ]")
    failed_list = export_existing_files(remote_path, before_path, commit_list)
    print(f"- 👌 원격 서버에서 {len(commit_list) - len(failed_list)}개의 파일을 성공적으로 백업했습니다.")
    print(f"- 😭 원격 서버에서 {len(failed_list)}개의 파일을 찾지 못했습니다.")

    print("[ 커밋을 진행합니다... ]")
    run_svn_commit(repo_path, commit_list, commit_message)
    print(f"- 🚀 {len(commit_list)}개의 파일이 성공적으로 커밋되었습니다.")

    print("\n[ 파일을 복사합니다... ]") 
    copy_changed_files(commit_list, repo_path, after_path)
    print(f"📦 {len(commit_list)}개의 파일을 '{after_path}'에 성공적으로 복사하였습니다.\n")

    res = input("변경 파일을 개발 서버에 반영할까요? (Y/N): ")

    if res == 'N':
        print("프로그램을 종료합니다.")
        press_input()
        return
    
    print("\n[ FTP 서버에 로그인합니다... ]") 
    username = input("FTP Username: ")
    password = input("FTP password: ")

    for host in ftp_list:
        transfer_committed_files(host, username, password)

    history = input("나중에 이 백업을 기억하기 위한 짧은 설명을 작성해주세요: ")

    with open(os.path.join(dest_path, "history.txt"), "w", encoding="UTF-8") as f:
        f.writelines(f"설명: {history}\n")
        f.writelines(f"커밋 메시지: {commit_message}\n")
        f.writelines(f"커밋 일시: {now}\n\n")
        
        # 커밋 리스트 기록
        f.writelines(f"[커밋 목록 {len(commit_list)}건]\n")
        for type, file_path in commit_list:
            f.writelines(f"({type}) - {file_path}\n")
        
        # 백업 실패 목록 기록 
        f.writelines(f"\n[백업 실패 목록 {len(failed_list)}건 - (SVN에 없는 파일)]\n")
        for type, file_path in failed_list:
            f.writelines(f"({type}) - {file_path}\n")

    excel_file_path = os.path.join(backup_path, "commit_history.xlsx")
    write_to_excel(commit_list, repo_path, excel_file_path)

    print("✅ Done!")
"""

def main() -> None:
    # 프로그램 세팅 파일 로딩
    setting_path = AppFiles.SETTING_FILE
    print()

    if not setting_path.exists():
        print(f"[에러] 😅 설정 파일을 찾을 수 없습니다. ('{setting_path}')")
        print(f"[에러] 자세한 설정 방법은 README.md 파일을 참고해주세요.")
        CmdUtil.press_input()
        exit(0)
    
    with open(setting_path, "r") as f:
        data = yaml.safe_load(f)
    
    projects = []

    print("\n[프로젝트 목록]")
    print("=" * 110)
    for index, (key, value) in enumerate(data.items(), start=1):
        print(f"{index}. {key}")
        for k in value:
            print(f"\t- {k}: {value[k]}")
        print()
        projects.append((key, value))
    print("=" * 110)

    while True:
        select_num = int(input("작업할 프로젝트 번호를 입력해주세요: "))
        select_num -= 1

        if select_num < 0 or select_num >= len(projects):
            print("[경고] 잘못된 프로젝트 번호입니다. 다시 선택해주세요.")
        else:
            break
    
    project = projects[select_num]
 
    if not project:
        print("\n[에러] 선택한 프로젝트의 설정값을 확인해주세요.")

        if 'local' not in project:
            print(f"- [에러] Local 경로를 지정해주세요.")
        if 'remote' not in project:
            print(f"- [에러] Remote 경로를 지정해주세요.")
        if 'ftp' not in project:
            print(f"- [에러] FTP 경로를 지정해주세요.")

        CmdUtil.press_input()
        exit(0)

    # Run
    svn_tracker = SvnTracker(project)
    svn_tracker.run()

    CmdUtil.press_input()


if __name__ == "__main__":
    AUTHOR = "sshwang"
    VERSION = "v1.0.3"

    CmdUtil.clear_cmd()

    print(f"""
    _____  _   _  _   _   _____                     _                
   /  ___|| | | || \\ | | |_   _|                   | |               
   \\ `--. | | | ||  \\| |   | |   _ __   __ _   ___ | | __  ___  _ __ 
    `--. \\| | | || . ` |   | |  | '__| / _` | / __|| |/ / / _ \\| '__|
   /\\__/ /\\ \\_/ /| |\\  |   | |  | |   | (_| || (__ |   < |  __/| |   
   \\____/  \\___/ \\_| \\_/   \\_/  |_|    \\__,_| \\___||_|\\_\\ \\___||_|   

                                    - developed by {AUTHOR} ({VERSION})
    """)

    main()