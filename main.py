import subprocess
import datetime
import os
import shutil
import yaml
from openpyxl import Workbook, load_workbook
from datetime import datetime
from src.utils import mkdir, svn_status_fullname

BACKUP_FOLDER_NAME = "SVN Tracker"
EXCEL_FILE_NAME = "history.xlsx"
EXCLUDE_FILE_TYPE = [".jar"]
EXCLUDE_FOLDER_TYPE = [".idea", "target", "lib", ".smarttomcat"]    

def run_svn_commit(repo_path, commit_list, commit_message):
    try:
        commit_cmd = ["svn", "commit", "-m", commit_message]

        for _, file_name in commit_list:
            commit_cmd.append(file_name)

        subprocess.run(commit_cmd, cwd=repo_path, check=True)
    except subprocess.CalledProcessError as e:
        print("SVN 커밋에 실패하였습니다.")
        print(e)
        exit(1)

def get_last_author(file_path, repo_path):
    try:
        result = subprocess.run(["svn", "log", "-l", "1", file_path], cwd=repo_path, capture_output=True, text=True)
        lines = result.stdout.splitlines()
        if len(lines) >= 2:
            parts = lines[1].split('|')
            if len(parts) >= 2:
                return parts[1].strip()
    except Exception:
        pass
    return "Unknown"

def copy_changed_files(commit_list, src_root, dest_root):
    # dest 경로가 없으면 폴더 생성
    if not os.path.exists(dest_root):
        os.makedirs(dest_root)

    for _, file_path in commit_list:
        full_src_path = os.path.join(src_root, file_path)
        full_dest_path = os.path.join(dest_root, file_path)

        os.makedirs(os.path.dirname(full_dest_path), exist_ok=True)

        if os.path.isfile(full_src_path):
            shutil.copy2(full_src_path, full_dest_path)
            print(f"- from {full_src_path}")
            print(f"- to {full_dest_path}")

def write_to_excel(changes, repo_path, excel_path):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 파일이 없으면 새로 생성
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Author", "Change Type", "File Path"])
    else:
        wb = load_workbook(excel_path)
        ws = wb.active

    for change_type, file_path in changes:
        author = get_last_author(file_path, repo_path)
        ws.append([timestamp, author, change_type, file_path])

    wb.save(excel_path)

def export_existing_files(remote_svn_path, before_path, commit_list):
    failed_list = []

    for type, file_path in commit_list:
        remote_file_path = remote_svn_path + "/" + file_path
        local_save_path = before_path + "/" + file_path

        try:
            if not os.path.exists(local_save_path):
                os.makedirs(local_save_path, exist_ok=True)

            subprocess.run(["svn", "export", remote_file_path, local_save_path], check=True)
        except subprocess.CalledProcessError as e:
            print("😭😭 {remote_file_path} 파일 다운로드 실패")
            print(e)
            failed_list.append((svn_status_fullname(type), file_path))
    
    return failed_list

def get_svn_status(repo_path):
    result = subprocess.run(["svn", "status"], cwd=repo_path, capture_output=True, text=True)
    return result.stdout

def parse_status_output(output):
    print(output)
    # A	추가됨 (Added)
    # C	충돌 (Conflicted)
    # D	삭제됨 (Deleted)
    # I	무시됨 (Ignored)
    # M	수정됨 (Modified)
    # R	교체됨 (Replaced)
    # X	외부 참조 (eXternals definition)
    # ?	버전 관리되지 않음 (Unversioned)
    # !	누락됨 (item missing, may be deleted)
    # ~	타입 변경됨 (Type changed)

    # 복합 상태가 나오는 경우
    # A +	(A: Added, +: with history)	파일을 추가했는데, 기존 파일 복사해서 만든 경우
    # R +	(R: Replaced, +: with history)	기존 파일 삭제 후 새 파일을 추가한 경우, 히스토리 있음
    # R C	(R: Replaced, C: Conflict)	파일을 교체했는데 충돌도 남
    # M C	(M: Modified, C: Conflict)	수정 중 충돌 발생
    # A C	(A: Added, C: Conflict)	추가한 파일에 충돌 발생 (드물지만 가능)

    BASE_STATUS = ['A', 'C', 'D', 'I', 'M', 'R', 'X', '?', '!', '~', 'A']
    changes = []

    for line in output.splitlines():
        if not line.strip():
            continue

        status_code = line[:7].strip()
        path = line[7:].strip()

        # 기본 상태인 경우
        if status_code in BASE_STATUS:
            # 충돌 파일 제외
            if status_code == 'C':
                print("- [Exclude] 충돌 파일 제외")
                print(line)
                print()
                continue

            # Unversioned file 제외
            if status_code == '?':
                print(f"- [Exclude] {path}  --  Unversioned File")
                continue
       
            print(f"- [Include] {path}")
            path = path.replace("\\", "/")
            status_code = svn_status_fullname(status_code)
            changes.append((status_code, path))

    return changes

def main(data: dict[str, object]) -> None:
    projects = []

    print("\n[Loading project list]\n")
    for index, (key, value) in enumerate(data.items(), start=1):
        print(f"{index}. {key}")
        
        for k in value:
            print(f"\t- {k}: {value[k]}")

        print("\n")
        projects.append(value)
    print("=" * 110)

    select_num = int(input("Enter the project number you want to select: "))
    select_num -= 1

    if select_num < 0 or select_num >= len(projects):
        print("Invalid project number selected.")
        print(f"Selected number {select_num + 1} does not exist in the project list.")
        exit(0)

    repo_path = projects[select_num]['local']
    remote_path = projects[select_num]['remote']

    if remote_path is None or repo_path is None:
        print("Error: Some settings are not configured.")
        print(f"- Remote Path: {remote_path}")
        print(f"- Local Path: {repo_path}")
        exit(1)

    repo_path = os.path.abspath(repo_path)
    repo_name = os.path.basename(repo_path)
    desktop_one_drive_path = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    if os.path.exists(desktop_one_drive_path):
        backup_path = os.path.join(desktop_one_drive_path, BACKUP_FOLDER_NAME, repo_name)
    elif os.path.exists(desktop_path):
        backup_path = os.path.join(desktop_path, BACKUP_FOLDER_NAME, repo_name)

    mkdir(backup_path)

    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest_path = os.path.join(backup_path, now)
    before_path = os.path.join(backup_path, now, "before")
    after_path = os.path.join(backup_path, now, "after")

    for path in [before_path, after_path]:
        mkdir(path)

    print("=" * 110)
    print("\n\n[ SVN STATUS ]") 
    print(f"👌 Target Repository Name: {repo_name}")
    print(f"📁 Repository Path: {repo_path}")
    print(f"📁 Backup Path: {dest_path}")
    print("\n🔍 Checking for changes...")
    status_output = get_svn_status(repo_path)

    changes = parse_status_output(status_output)
    print("\n")

    if not changes:
        print("✅ No changes detected.")
        return

    commit_list = []
    print("[ Select files to commit ]")

    for type, file_path in changes:
        ans = input(f"({type}) {file_path} (Y/N): ")
        if ans.upper() == "Y":
            commit_list.append((type, file_path))

    if len(commit_list) == 0:
        print("✅ No files selected for commit.")
        return

    print("\n\n[ Files to commit ]")
    for type, file_path in commit_list:
        print(f"- ({type}) {file_path}")

    rst = input("\n\nProceed with commit? (Y/N): ")

    if rst.upper() != "Y":
        return

    while True:
        commit_message = input("Enter commit message: ")
        if len(commit_message) > 0:
            break
        else:
            print("- Commit message is required. Please describe the changes.")

    print("\n\n[ Downloading existing files ]")
    failed_list = export_existing_files(remote_path, before_path, commit_list)
    print(f"- 👌 Successfully download {len(commit_list) - len(failed_list)} files.")
    print(f"- 😭 Failed to download {len(failed_list)} files. (Not found in the remote SVN repository)")

    print("[ Committing files.... ]")
    run_svn_commit(repo_path, commit_list, commit_message)
    print(f"- 🚀 {len(commit_list)} files have been committed successfully.")

    print("\n\n[ COPYING FILES.... ]") 
    print(f"📦 Copying {len(commit_list)} changed files to {after_path}")
    copy_changed_files(commit_list, repo_path, after_path)
    print()

    history = input("Write a brief history description for this backup: ")

    with open(os.path.join(dest_path, "history.txt"), "w", encoding="UTF-8") as f:
        f.writelines(f"Description: {history}\n")
        f.writelines(f"Commit Message: {commit_message}\n")
        f.writelines(f"Commit Date: {now}\n\n")
        
        # 커밋 리스트 기록
        f.writelines(f"[커밋 목록 {len(commit_list)}건]\n")
        for type, file_path in commit_list:
            f.writelines(f"({type}) - {file_path}\n")
        
        # 백업 실패 목록 기록 
        f.writelines(f"\n[백업 실패 목록 {len(failed_list)}건 - (SVN에 없는 파일)]\n")
        for type, file_path in failed_list:
            f.writelines(f"({type}) - {file_path}\n")

    excel_file_path = os.path.join(backup_path, "commit_history.xlsx")
    write_to_excel(commit_list, repo_path, excel_file_path)

    print("✅ Done!")

if __name__ == "__main__":
    print("""
    _____  _   _  _   _   _____                     _                
   /  ___|| | | || \\ | | |_   _|                   | |               
   \\ `--. | | | ||  \\| |   | |   _ __   __ _   ___ | | __  ___  _ __ 
    `--. \\| | | || . ` |   | |  | '__| / _` | / __|| |/ / / _ \\| '__|
   /\\__/ /\\ \\_/ /| |\\  |   | |  | |   | (_| || (__ |   < |  __/| |   
   \\____/  \\___/ \\_| \\_/   \\_/  |_|    \\__,_| \\___||_|\\_\\ \\___||_|   

                                            - developed by sshwang
    """)

    # 프로그램 세팅 파일 로딩
    setting_path = os.path.join(os.path.expanduser("~"), "svn_tracker.yaml")

    if not os.path.exists(setting_path):
        print(f"😅 Could not find {setting_path}\n\n")
    else:
        with open(setting_path, "r") as yml:
            data = yaml.safe_load(yml)
        
        main(data)

    input("Press Enter to exit...")