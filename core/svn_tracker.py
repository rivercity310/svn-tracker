import subprocess
import shutil
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime
from pathlib import Path
from core.path_manager import PathManager
from utils.cmd_util import CmdUtil
from clients.ssh_client import SshClient


# A	추가됨 (Added)
# C	충돌 (Conflicted)
# D	삭제됨 (Deleted)
# I	무시됨 (Ignored)
# M	수정됨 (Modified)
# R	교체됨 (Replaced)
# X	외부 참조 (eXternals definition)
# ?	버전 관리되지 않음 (Unversioned)
# !	누락됨 (item missing, may be deleted)
# ~	타입 변경됨 (Type changed)

# 복합 상태가 나오는 경우
# A +	(A: Added, +: with history)	파일을 추가했는데, 기존 파일 복사해서 만든 경우
# R +	(R: Replaced, +: with history)	기존 파일 삭제 후 새 파일을 추가한 경우, 히스토리 있음
# R C	(R: Replaced, C: Conflict)	파일을 교체했는데 충돌도 남
# M C	(M: Modified, C: Conflict)	수정 중 충돌 발생
# A C	(A: Added, C: Conflict)	추가한 파일에 충돌 발생 (드물지만 가능)
class SvnTracker:
    STATUS_MAP = {
        'A': 'Added', 'C': 'Conflicted', 'D': 'Deleted', 'I': 'Ignored',
        'M': 'Modified', 'R': 'Replaced', 'X': 'External', '?': 'Unversioned',
        '!': 'Missing', '~': 'Obstructed'
    }

    def __init__(self, project: tuple[str, any]) -> None:
        self.project_name, config = project
        self.local, self.remote, self.ssh = config['local'], config['remote'], config['ssh']
        self.now = datetime.now().strftime("%Y%m%d_%H%M")
        self.workdir = PathManager.get_workdir(self.project_name) / self.now
        self.workdir_before = self.workdir / "before"
        self.workdir_after = self.workdir / "after"

        # 경로 초기화 (이전, 이후)
        PathManager.mkdir(self.workdir_before)
        PathManager.mkdir(self.workdir_after)


    def run(self) -> None:
        print("=" * 110, "\n\n[ SVN STATUS ]")
        print(f"👌 Local Repository Path: {self.workdir}\n📁 Remote Repository Path: {self.remote}\n")
        print("🔍 Checking for changes...")

        status_output = self._get_svn_status()
        print(status_output)

        change_files = self._parse_svn_status(status_output)
        if not change_files:
            print("✅ 변경된 파일이 없습니다.\n프로그램을 종료합니다.")
            CmdUtil.press_input()
            return

        # 커밋 대상 파일 선정
        print("\n\n[ 커밋할 대상 파일을 선택해주세요. ]")
        commit_list = [
            {'type': type_, 'path': path}
            for type_, path in change_files if input(f"({type_}) {path} (Y/N): ").strip().upper() == 'Y'
        ]

        if len(commit_list) == 0:
            print("✅ 선택된 커밋 대상 파일이 없습니다.\n프로그램을 종료합니다.")
            CmdUtil.press_input()
            return

        print("\n\n[ 선택된 커밋 대상 파일 ]")
        for c in commit_list: print(f"\t- ({c['type']}) {c['path']}")

        if input("\n커밋을 진행할까요? (Y/N): ").strip().upper() != 'Y':
            print("프로그램을 종료합니다.")
            CmdUtil.press_input()
            return

        commit_message = input("커밋 메시지를 입력해주세요: ").strip()
        while not commit_message:
            commit_message = input("[경고] 커밋 메시지는 최소 1글자 이상 입력해야 합니다: ").strip()

        print("\n\n[ Remote 서버의 파일을 백업합니다. ]")
        success_list, fail_list = self._export_existing_files(commit_list)
        for s in success_list: print(f"\t- 👌 {s['status']}: {s['path']}")
        for f in fail_list: print(f"\t- 😭 {f['status']}: {f['path']}")

        print("\n\n[ 커밋을 진행합니다... ]")
        self._svn_commit(commit_list, commit_message)
        print(f"\t- 🚀 {len(commit_list)}개의 파일이 성공적으로 커밋되었습니다.")

        print("\n\n[ 파일을 복사합니다... ]") 
        self._copy_changed_files(commit_list)
        print(f"\t📦 {len(commit_list)}개의 파일을 '{self.workdir_after}'에 성공적으로 복사하였습니다.\n")

        print("\n\n[ 히스토리를 기록하고 있습니다... ]") 
        self._write_summary(commit_list, success_list, fail_list, commit_message)
        self._write_to_excel(commit_list, commit_message)
        print("✅ Done!")

        if input("\n변경 파일을 개발 서버에 반영할까요? (Y/N): ").strip().upper() != 'Y':
            print("프로그램을 종료합니다.")
            CmdUtil.press_input()
            return
       
        print("\n\n[ SSH 접속중... ]")
        ssh = SshClient(self.ssh, self.project_name)
        ssh.svn_update(commit_list)

        path_list = [c['path'].endswith('.java') or c['path'].endswith('.xml') for c in commit_list]

        if any(path_list):
            ssh.build_jar()
            ssh.tdown()
            time.sleep(7)
            ssh.tboot()

        print("✅ Done!")


    def _get_svn_status(self) -> str:
        result = subprocess.run(["svn", "status"], cwd=self.local, capture_output=True, text=True)
        return result.stdout


    def _parse_svn_status(self, status_output: str):
        changes = []

        for line in status_output.splitlines():
            status_code, path = line[:7].strip(), line[7:].strip().replace("\\", "/")

            if status_code == "C":
                print("- [Exclude] 충돌 파일 제외\n", line)
                continue

            if status_code == '?':
                print(f"- [Exclude] {path}  --  Unversioned File")
                continue

            if status_code in self.STATUS_MAP:
                print(f"- [Include] {path}")
                changes.append((self.STATUS_MAP.get(status_code, "Unknown"), path))

        return changes


    # 기존 파일 다운로드 (백업)
    def _export_existing_files(self, commit_list: list[dict]) -> tuple[list[dict], list[dict]]:
        success, fail = [], []

        for commit in commit_list:
            remote_file_path = f"{self.remote}/{commit['path']}"
            local_save_path = str(self.workdir_before / commit['path'])

            try:
                PathManager.mkdir(local_save_path)
                subprocess.run(["svn", "export", remote_file_path, local_save_path], check=True)
                success.append({"status": commit['type'], "path": commit['path']})
            except subprocess.CalledProcessError as e:
                print(f"\t😭😭 {remote_file_path} 파일 다운로드 실패")
                print(e)
                fail.append({"status": commit['type'], "path": commit['path']})

        return success, fail   
    

    def _svn_commit(self, commit_list: list, commit_message: str):
        try:
            subprocess.run(["svn", "commit", "-m", commit_message, *[c['path'] for c in commit_list]], cwd=self.local, check=True)
        except subprocess.CalledProcessError as e:
            print("SVN 커밋에 실패하였습니다.")
            print(e)
            exit(1)


    def _copy_changed_files(self, commit_list: list):
        for commit in commit_list:
            src, dest = Path(self.local) / commit['path'], self.workdir_after / commit['path']
            PathManager.mkdir(dest)

            if src.is_file():
                shutil.copy2(src, dest)
                print(f"- {src} ===> {dest}")


    def _write_to_excel(self, commit_list, commit_message):
        excel_file_path = self.workdir.parent / "commit_history.xlsx"
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 파일이 없으면 새로 생성
        if not excel_file_path.exists():
            wb = Workbook()
            ws = wb.active
            ws.append(["일시", "변경자", "변경 타입", "경로", "변경 사유"])
        else:
            wb = load_workbook(excel_file_path)
            ws = wb.active

        for commit in commit_list:
            file_path = commit['path']
            author = self._get_last_author(file_path)
            ws.append([timestamp, author, commit['type'], file_path, commit_message])

        wb.save(excel_file_path)


    def _get_last_author(self, file_path):
        try:
            result = subprocess.run(["svn", "log", "-l", "1", file_path], cwd=self.local, capture_output=True, text=True)
            return result.stdout.splitlines()[1].split('|')[1].strip()
        except Exception:
            return "Unknown"
        

    def _write_summary(self, commit_list, success_list, fail_list, commit_msg):
        with open(self.workdir / "summary.txt", "w", encoding="UTF-8") as f:
            f.write(f"커밋 메시지: {commit_msg}\n커밋 일시: {self.now}\n\n")

            f.write(f"[커밋 목록 {len(commit_list)}건]\n")
            for c in commit_list:
                f.write(f"({c['type']}) - {c['path']}\n")

            f.write(f"\n[백업 성공 목록 {len(success_list)}건]\n")
            for s in success_list:
                f.write(f"({s['status']}) - {s['path']}\n")

            f.write(f"\n[백업 실패 목록 {len(fail_list)}건 - (SVN에 없는 파일)]\n")
            for f_item in fail_list:
                f.write(f"({f_item['status']}) - {f_item['path']}\n")
